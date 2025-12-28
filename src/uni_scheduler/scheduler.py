from __future__ import annotations
import math
from collections import defaultdict
from typing import Dict, List, Optional, Tuple
from .models import Room, Professor, Course, Slot, Assignment, DAYS, minutes_to_hhmm, END_MINUTES
from .utils import generar_candidatos, profesor_disponible

class Scheduler:
    def __init__(self, semester: str, rooms: List[Room], professors: List[Professor], courses: List[Course], assistants_pool: List[str] = None):
        self.semester = semester  # "Abril-Agosto" | "Agosto-Diciembre"
        self.rooms = rooms
        self.professors: Dict[str, Professor] = {p.id: p for p in professors}
        self.courses = courses
        self.assistants_pool = assistants_pool or []
        self.room_occupancy: Dict[str, List[Assignment]] = defaultdict(list)
        self.prof_occupancy: Dict[str, List[Assignment]] = defaultdict(list)
        self.assignments: List[Assignment] = []
        self.alerts: List[str] = []

    # --- reglas de semestre ---
    def ciclo_permitido(self, cycle: int) -> bool:
        if self.semester == "Abril-Agosto":
            return cycle % 2 == 1
        if self.semester == "Agosto-Diciembre":
            return cycle % 2 == 0
        return True

    # --- helpers de ocupación ---
    def libre_room(self, room_id: str, slot: Slot) -> bool:
        return all(not slot.overlaps(a.slot) for a in self.room_occupancy[room_id])

    def libre_prof(self, prof_id: str, slot: Slot) -> bool:
        return all(not slot.overlaps(a.slot) for a in self.prof_occupancy[prof_id])

    def horas_prof_en_dia(self, prof_id: str, day: str) -> float:
        return sum(a.slot.duration_hours() for a in self.prof_occupancy[prof_id] if a.slot.day == day)

    # --- salas ---
    def salas_disponibles(self, faculty: str, kind: str, capacidad_req: int, slot: Slot) -> List[Room]:
        cands = [r for r in self.rooms if r.faculty == faculty and r.kind == kind and r.capacity >= capacidad_req]
        return [r for r in cands if self.libre_room(r.id, slot)]

    # --- colocar ---
    def colocar(self, course: Course, group: str, slot: Slot, room: Room, professor_id: Optional[str]):
        asg = Assignment(course.code, group, slot, room.id, professor_id)
        self.assignments.append(asg)
        self.room_occupancy[room.id].append(asg)
        if professor_id:
            self.prof_occupancy[professor_id].append(asg)

    # --- Fase 1: Labs primero ---
    def fase_labs(self):
        lab_courses = [c for c in self.courses if c.inscritos_lab > 0 and self.ciclo_permitido(c.cycle)]
        lab_courses.sort(key=lambda c: -math.ceil(c.inscritos_lab / 15))
        for c in lab_courses:
            grupos = math.ceil(c.inscritos_lab / 15)
            prof = self.professors.get(c.profesor_id) if c.profesor_id else None
            slots = generar_candidatos(c.duracion_lab_horas)
            for g in range(1, grupos + 1):
                asignado = False
                for slot in slots:
                    salas = self.salas_disponibles(c.faculty, "lab", 15, slot)
                    if not salas:
                        continue
                    prof_id = None
                    if prof and prof.disponible_labs and profesor_disponible(prof, slot) and self.libre_prof(prof.id, slot):
                        if self.horas_prof_en_dia(prof.id, slot.day) + slot.duration_hours() <= 6:
                            prof_id = prof.id
                    # si el prof no hace labs, se permite asistente (prof_id None)
                    self.colocar(c, f"Lab-{g}", slot, salas[0], prof_id)
                    asignado = True
                    break
                if not asignado:
                    self.alerts.append(f"Laboratorio no asignado: {c.name} requiere grupo Lab-{g} y no hay slot disponible")

    # --- Fase 2: Teóricos ---
    def fase_teoricos(self):
        teo_courses = [c for c in self.courses if c.inscritos_teorico > 0 and self.ciclo_permitido(c.cycle)]
        for c in teo_courses:
            grupos = []
            if c.inscritos_teorico > 60:
                grupos.append(("Teo-A", 60))
                grupos.append(("Teo-B", c.inscritos_teorico - 60))
            else:
                grupos.append(("Teo-A", c.inscritos_teorico))

            # proximidad a labs: día medio
            lab_days_idx = [DAYS.index(a.slot.day) for a in self.assignments if a.course_code == c.code and a.group.startswith("Lab-")]
            target_day = None
            if lab_days_idx:
                avg = round(sum(lab_days_idx) / len(lab_days_idx))
                target_day = DAYS[min(max(avg, 0), len(DAYS)-1)]
            slots = generar_candidatos(c.duracion_teorico_horas)
            if target_day:
                slots.sort(key=lambda s: abs(DAYS.index(s.day) - DAYS.index(target_day)))

            prof = self.professors.get(c.profesor_id) if c.profesor_id else None
            if prof and prof.habilitado_desde_ciclo > c.cycle:
                self.alerts.append(f"Profesor no habilitado: {prof.name} desde {prof.habilitado_desde_ciclo}+ para {c.name} (ciclo {c.cycle})")
                prof = None

            if len(grupos) == 2:
                if not self._colocar_teorico_doble_consecutivo(c, grupos, slots, prof):
                    for (gname, cap) in grupos:
                        self._colocar_teorico_simple(c, gname, cap, slots, prof)
            else:
                gname, cap = grupos[0]
                self._colocar_teorico_simple(c, gname, cap, slots, prof)

    def _colocar_teorico_doble_consecutivo(self, c: Course, grupos, slots, prof: Optional[Professor]) -> bool:
        dur = c.duracion_teorico_horas * 60
        gap = 30
        for s in slots:
            s2 = Slot(s.day, s.end + gap, s.end + gap + dur)
            if s2.end > END_MINUTES:
                continue
            salas_A = self.salas_disponibles(c.faculty, "teorico", min(60, grupos[0][1]), s)
            salas_B = self.salas_disponibles(c.faculty, "teorico", min(60, grupos[1][1]), s2)
            if not salas_A or not salas_B:
                continue
            prof_id = None
            if prof and all([profesor_disponible(prof, s), profesor_disponible(prof, s2), self.libre_prof(prof.id, s), self.libre_prof(prof.id, s2)]):
                if self.horas_prof_en_dia(prof.id, s.day) + s.duration_hours() + s2.duration_hours() <= 6:
                    prof_id = prof.id
            self.colocar(c, grupos[0][0], s, salas_A[0], prof_id)
            self.colocar(c, grupos[1][0], s2, salas_B[0], prof_id)
            return True
        return False

    def _colocar_teorico_simple(self, c: Course, gname: str, cap: int, slots, prof: Optional[Professor]):
        for s in slots:
            salas = self.salas_disponibles(c.faculty, "teorico", min(60, cap), s)
            if not salas:
                continue
            prof_id = None
            if prof and profesor_disponible(prof, s) and self.libre_prof(prof.id, s):
                if self.horas_prof_en_dia(prof.id, s.day) + s.duration_hours() <= 6:
                    prof_id = prof.id
            elif prof:
                # profesor existe pero no cabe en disponibilidad/tope
                pass
            self.colocar(c, gname, s, salas[0], prof_id)
            return
        self.alerts.append(f"Teórico no asignado: {c.name} {gname}")

    # --- Fase 3: Validaciones ---
    def fase_validaciones(self):
        # cursos fuera de semestre
        for c in self.courses:
            if not self.ciclo_permitido(c.cycle):
                self.alerts.append(f"Curso fuera del ciclo permitido ({self.semester}): {c.name} (ciclo {c.cycle})")

        # huecos > 2 días entre teórico y lab
        for c in self.courses:
            dias_teo = [a for a in self.assignments if a.course_code == c.code and a.group.startswith("Teo-")]
            dias_lab = [a for a in self.assignments if a.course_code == c.code and a.group.startswith("Lab-")]
            if dias_teo and dias_lab:
                t_idx = [DAYS.index(a.slot.day) for a in dias_teo]
                l_idx = [DAYS.index(a.slot.day) for a in dias_lab]
                min_gap = min(abs(t - l) for t in t_idx for l in l_idx)
                if min_gap > 2:
                    self.alerts.append(f"Hueco >2 días entre teórico y lab: {c.name} (min {min_gap} días)")

        # profesor >6h por día
        for pid, asigns in self.prof_occupancy.items():
            horas = {}
            for a in asigns:
                horas.setdefault(a.slot.day, 0.0)
                horas[a.slot.day] += a.slot.duration_hours()
            for d, h in horas.items():
                if h > 6:
                    self.alerts.append(f"Profesor {self.professors[pid].name} sobrepasó horas máximas el {d}: {h:.1f}h")

    def build(self):
        self.fase_labs()
        self.fase_teoricos()
        self.fase_validaciones()

    # --- Export ---
    def export_excel(self, path: str = "horarios.xlsx"):
        try:
            from openpyxl import Workbook
        except Exception:
            return
        wb = Workbook()
        wb.remove(wb.active)
        por_fac_ciclo = defaultdict(list)
        cursos_map = {c.code: c for c in self.courses}
        for a in self.assignments:
            c = cursos_map[a.course_code]
            por_fac_ciclo[(c.faculty, c.cycle)].append(a)
        for (fac, cyc), asigns in por_fac_ciclo.items():
            title = f"{fac}-C{cyc}"[:31]
            ws = wb.create_sheet(title)
            ws.append(["Curso", "Grupo", "Día", "Inicio", "Fin", "Sala", "Profesor"])
            for a in sorted(asigns, key=lambda x: (DAYS.index(x.slot.day), x.slot.start, x.course_code, x.group)):
                prof_name = self.professors[a.professor_id].name if a.professor_id else "Asistente/No asignado"
                ws.append([cursos_map[a.course_code].name, a.group, a.slot.day, minutes_to_hhmm(a.slot.start), minutes_to_hhmm(a.slot.end), a.room_id, prof_name])
        ws = wb.create_sheet("Alertas")
        ws.append(["Mensaje"])
        for msg in self.alerts:
            ws.append([msg])
        wb.save(path)

    def export_pdf_alertas(self, path: str = "alertas.pdf"):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
        except Exception:
            return
        c = canvas.Canvas(path, pagesize=A4)
        width, height = A4
        x, y = 40, height - 50
        c.setFont("Helvetica-Bold", 14)
        c.drawString(x, y, "Alertas de Horarios")
        y -= 20
        c.setFont("Helvetica", 10)
        if not self.alerts:
            c.drawString(x, y, "No se registraron alertas.")
        else:
            for msg in self.alerts:
                if y < 60:
                    c.showPage()
                    y = height - 50
                    c.setFont("Helvetica", 10)
                c.drawString(x, y, f"• {msg}")
                y -= 14
        c.save()
